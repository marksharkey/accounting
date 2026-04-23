"""Import Chase Checking transactions from QBO export."""
import xlrd
from decimal import Decimal
from datetime import datetime
import sys
from database import SessionLocal
import models

def parse_qbo_date(date_value):
    """Parse date from Excel (could be float or string)."""
    if isinstance(date_value, float):
        # Excel date format (days since 1900-01-01)
        import openpyxl
        from openpyxl.utils import get_column_letter
        # Try to parse as Excel date
        try:
            from datetime import timedelta, date as dt
            excel_epoch = dt(1900, 1, 1)
            return excel_epoch + timedelta(days=int(date_value) - 2)
        except:
            pass

    # Try string parsing
    if isinstance(date_value, str):
        try:
            return datetime.strptime(date_value, '%m/%d/%Y').date()
        except:
            pass

    return None

def map_qbo_type_to_transaction_type(qbo_type):
    """Map QBO transaction type to our TransactionType enum."""
    type_map = {
        'Check': 'check',
        'Journal': 'other',
        'Payment': 'payment',
    }
    return type_map.get(qbo_type, 'other')

def import_chase_checking():
    """Import Chase Checking register from QBO export."""
    # Open the Excel file
    wb = xlrd.open_workbook('/Users/marksharkey/Downloads/Register.xls')
    ws = wb.sheet_by_index(0)

    db = SessionLocal()

    # Get or create the Chase Checking bank account
    account = db.query(models.BankAccount).filter_by(account_name='Chase Checking').first()
    if not account:
        account = models.BankAccount(
            account_name='Chase Checking',
            account_number='****',
            account_type='Checking',
            opening_balance=Decimal('0.00'),
            is_active=True
        )
        db.add(account)
        db.commit()
        print(f"Created bank account: {account.account_name}")

    # Delete existing transactions for this account to start fresh
    db.query(models.BankTransaction).filter_by(bank_account_id=account.id).delete()
    db.commit()
    print(f"Cleared existing transactions for {account.account_name}")

    # Import transactions
    imported_count = 0
    skipped_count = 0

    # Skip header rows (0 and 1)
    for row_idx in range(2, ws.nrows):
        try:
            # Extract values
            date_val = ws.cell(row_idx, 0).value
            ref_no = ws.cell(row_idx, 1).value
            payee = ws.cell(row_idx, 2).value
            memo = ws.cell(row_idx, 3).value
            payment = ws.cell(row_idx, 4).value
            deposit = ws.cell(row_idx, 5).value
            reconciliation_status = ws.cell(row_idx, 6).value
            balance = ws.cell(row_idx, 7).value
            transaction_type = ws.cell(row_idx, 8).value
            gl_account = ws.cell(row_idx, 9).value

            # Parse date
            transaction_date = parse_qbo_date(date_val)
            if not transaction_date:
                skipped_count += 1
                continue

            # Determine amount and direction
            if payment:
                amount = Decimal(str(payment)) * -1  # Negative for payments
            elif deposit:
                amount = Decimal(str(deposit))  # Positive for deposits
            else:
                skipped_count += 1
                continue

            # Create description from payee and memo
            description_parts = []
            if payee and payee.strip():
                description_parts.append(str(payee))
            if memo and memo.strip():
                description_parts.append(str(memo))
            description = ' - '.join(description_parts) if description_parts else ''

            # Handle transaction number - could be None
            transaction_number = None
            if ref_no:
                transaction_number = str(ref_no).strip() if isinstance(ref_no, str) else str(int(ref_no)) if isinstance(ref_no, float) else None

            # Create transaction
            txn = models.BankTransaction(
                bank_account_id=account.id,
                transaction_date=transaction_date,
                transaction_type=map_qbo_type_to_transaction_type(transaction_type),
                transaction_number=transaction_number,
                description=description if description else None,
                gl_account=str(gl_account).strip() if gl_account else None,
                amount=amount,
                balance=Decimal(str(balance)) if balance else None,
                reconciled=reconciliation_status == 'Reconciled',
                import_batch='qbo_export_2026_04_18'
            )
            db.add(txn)
            imported_count += 1

            if imported_count % 100 == 0:
                db.commit()
                print(f"  Imported {imported_count} transactions...")

        except Exception as e:
            skipped_count += 1
            print(f"  Error on row {row_idx}: {e}")
            continue

    db.commit()
    print(f"\nImport complete!")
    print(f"  Imported: {imported_count}")
    print(f"  Skipped: {skipped_count}")
    db.close()

if __name__ == '__main__':
    import_chase_checking()
